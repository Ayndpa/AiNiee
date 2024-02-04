import datetime
import json
import random
import os
from PyQt5.QtWidgets import QFileDialog
import openpyxl
from ..Global import Global


# 文件读取器
class File_Reader:
    def __init__(self):
        pass

    # 选择输入文件夹按钮绑定函数
    def Select_project_folder(self):
        Input_Folder = QFileDialog.getExistingDirectory(
            None, "Select Directory", ""
        )  # 调用QFileDialog类里的函数来选择文件目录
        if Input_Folder:
            # 将输入路径存储到配置器中
            Global.configurator.Input_Folder = Input_Folder
            Global.window.Widget_translation_settings.A_settings.label_input_path.setText(
                Input_Folder
            )
            print("[INFO]  已选择项目文件夹: ", Input_Folder)
        else:
            print("[INFO]  未选择文件夹")
            return  # 直接返回，不执行后续操作

    # 选择输入文件夹按钮绑定函数(检查任务用)
    def Select_project_folder_check(self):
        Input_Folder = QFileDialog.getExistingDirectory(
            None, "Select Directory", ""
        )  # 调用QFileDialog类里的函数来选择文件目录
        if Input_Folder:
            # 将输入路径存储到配置器中
            Global.configurator.Input_Folder = Input_Folder
            Global.window.Widget_check.label_input_path.setText(Input_Folder)
            print("[INFO]  已选择项目文件夹: ", Input_Folder)
        else:
            print("[INFO]  未选择文件夹")
            return  # 直接返回，不执行后续操作

    # 选择输出文件夹按钮绑定函数
    def Select_output_folder(self):
        Output_Folder = QFileDialog.getExistingDirectory(
            None, "Select Directory", ""
        )  # 调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            # 将输入路径存储到配置器中
            Global.configurator.Output_Folder = Output_Folder
            Global.window.Widget_translation_settings.A_settings.label_output_path.setText(
                Output_Folder
            )
            print("[INFO]  已选择输出文件夹:", Output_Folder)
        else:
            print("[INFO]  未选择文件夹")
            return  # 直接返回，不执行后续操作

    # 选择输出文件夹按钮绑定函数(检查任务用)
    def Select_output_folder_check(self):
        Output_Folder = QFileDialog.getExistingDirectory(
            None, "Select Directory", ""
        )  # 调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            # 将输入路径存储到配置器中
            Global.configurator.Output_Folder = Output_Folder
            Global.window.Widget_check.label_output_path.setText(Output_Folder)
            print("[INFO]  已选择输出文件夹:", Output_Folder)
        else:
            print("[INFO]  未选择文件夹")
            return  # 直接返回，不执行后续操作

    # 生成项目ID
    def generate_project_id(self, prefix):
        # 获取当前时间，并将其格式化为数字字符串
        current_time = datetime.datetime.now().strftime("%Y%m%d%H%M%S")

        # 生成5位随机数
        random_number = random.randint(10000, 99999)

        # 组合生成项目ID
        project_id = f"{current_time}{prefix}{random_number}"

        return project_id

    # 读取文件夹中树形结构json文件
    def read_mtool_files(self, folder_path):
        # 缓存数据结构示例
        ex_cache_data = [
            {"project_type": "Mtool"},
            {
                "text_index": 1,
                "text_classification": 0,
                "translation_status": 0,
                "source_text": "しこトラ！",
                "translated_text": "无",
                "storage_path": "TrsData.json",
                "file_name": "TrsData.json",
            },
            {
                "text_index": 2,
                "text_classification": 0,
                "translation_status": 0,
                "source_text": "室内カメラ",
                "translated_text": "无",
                "storage_path": "TrsData.json",
                "file_name": "TrsData.json",
            },
            {
                "text_index": 3,
                "text_classification": 0,
                "translation_status": 0,
                "source_text": "室内カメラ",
                "translated_text": "无",
                "storage_path": "DEBUG Folder\\Replace the original text.json",
                "file_name": "Replace the original text.json",
            },
        ]

        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self, "Mtool")
        json_data_list.append(
            {
                "project_type": "Mtool",
                "project_id": project_id,
            }
        )

        # 文本索引初始值
        i = 1

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 JSON 文件
                if file.endswith(".json"):
                    file_path = os.path.join(root, file)  # 构建文件路径

                    # 读取 JSON 文件内容
                    with open(file_path, "r", encoding="utf-8") as json_file:
                        json_data = json.load(json_file)

                        # 提取键值对
                        for key, value in json_data.items():
                            # 根据 JSON 文件内容的数据结构，获取相应字段值
                            source_text = key
                            translated_text = value
                            storage_path = os.path.relpath(file_path, folder_path)
                            file_name = file
                            # 将数据存储在字典中
                            json_data_list.append(
                                {
                                    "text_index": i,
                                    "text_classification": 0,
                                    "translation_status": 0,
                                    "source_text": source_text,
                                    "translated_text": translated_text,
                                    "semantic_similarity": 0,
                                    "storage_path": storage_path,
                                    "file_name": file_name,
                                }
                            )

                            # 增加文本索引值
                            i = i + 1

        return json_data_list

    # 读取文件夹中树形结构的xlsx文件， 存到列表变量中
    def read_xlsx_files(self, folder_path):
        # 缓存数据结构示例
        ex_cache_data = [
            {"project_type": "T++"},
            {
                "text_index": 1,
                "text_classification": 0,
                "translation_status": 0,
                "source_text": "しこトラ！",
                "translated_text": "无",
                "storage_path": "TrsData.xlsx",
                "file_name": "TrsData.xlsx",
                "row_index": 1,
            },
            {
                "text_index": 2,
                "text_classification": 0,
                "translation_status": 0,
                "source_text": "室内カメラ",
                "translated_text": "无",
                "storage_path": "TrsData.xlsx",
                "file_name": "TrsData.xlsx",
                "row_index": 2,
            },
            {
                "text_index": 3,
                "text_classification": 0,
                "translation_status": 0,
                "source_text": "室内カメラ",
                "translated_text": "无",
                "storage_path": "DEBUG Folder\\text.xlsx",
                "file_name": "text.xlsx",
                "row_index": 3,
            },
        ]

        # 创建列表
        cache_list = []
        # 添加文件头
        project_id = File_Reader.generate_project_id(self, "T++")
        cache_list.append(
            {
                "project_type": "T++",
                "project_id": project_id,
            }
        )
        # 文本索引初始值
        i = 1

        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.endswith(".xlsx"):
                    file_path = os.path.join(root, file)  # 构建文件路径

                    wb = openpyxl.load_workbook(file_path)
                    sheet = wb.active
                    for row in range(
                        2, sheet.max_row + 1
                    ):  # 从第二行开始读取，因为第一行是标识头，通常不用理会
                        cell_value1 = sheet.cell(
                            row=row, column=1
                        ).value  # 第N行第一列的值
                        cell_value2 = sheet.cell(
                            row=row, column=2
                        ).value  # 第N行第二列的值

                        source_text = cell_value1  # 获取原文
                        storage_path = os.path.relpath(
                            file_path, folder_path
                        )  # 用文件的绝对路径和输入文件夹路径“相减”，获取相对的文件路径
                        file_name = file  # 获取文件名

                        # 第1列的值不为空，和第2列的值为空，是未翻译内容
                        if cell_value1 and cell_value2 is None:

                            translated_text = "无"
                            cache_list.append(
                                {
                                    "text_index": i,
                                    "text_classification": 0,
                                    "translation_status": 0,
                                    "source_text": source_text,
                                    "translated_text": translated_text,
                                    "semantic_similarity": 0,
                                    "storage_path": storage_path,
                                    "file_name": file_name,
                                    "row_index": row,
                                }
                            )

                            i = i + 1  # 增加文本索引值

                        # 第1列的值不为空，和第2列的值不为空，是已经翻译内容
                        elif cell_value1 and cell_value2:

                            translated_text = cell_value2
                            cache_list.append(
                                {
                                    "text_index": i,
                                    "text_classification": 0,
                                    "translation_status": 1,
                                    "source_text": source_text,
                                    "translated_text": translated_text,
                                    "storage_path": storage_path,
                                    "semantic_similarity": 0,
                                    "file_name": file_name,
                                    "row_index": row,
                                }
                            )

                            i = i + 1  # 增加文本索引值

        return cache_list

    # 读取缓存文件
    def read_cache_files(self, folder_path):
        # 获取文件夹中的所有文件
        files = os.listdir(folder_path)

        # 查找以 "CacheData" 开头且以 ".json" 结尾的文件
        json_files = [
            file
            for file in files
            if file.startswith("AinieeCacheData") and file.endswith(".json")
        ]

        if not json_files:
            print(f"Error: No 'CacheData' JSON files found in folder '{folder_path}'.")
            return None

        # 选择第一个符合条件的 JSON 文件
        json_file_path = os.path.join(folder_path, json_files[0])

        # 读取 JSON 文件内容
        with open(json_file_path, "r", encoding="utf-8") as json_file:
            data = json.load(json_file)
            return data

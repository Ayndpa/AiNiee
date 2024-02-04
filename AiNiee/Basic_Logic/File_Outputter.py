import copy
import json
import opencc
import os

from openpyxl import Workbook


# 文件输出器
class File_Outputter:
    def __init__(self):
        pass

    # 将缓存文件里已翻译的文本转换为简体字或繁体字
    def simplified_and_traditional_conversion(self, cache_list, target_language):
        # 缓存数据结构示例
        ex_cache_data = [
            {"project_type": "Mtool"},
            {
                "text_index": 1,
                "text_classification": 0,
                "translation_status": 1,
                "source_text": "しこトラ！",
                "translated_text": "谢谢",
                "storage_path": "TrsData.json",
                "file_name": "TrsData.json",
            },
            {
                "text_index": 2,
                "text_classification": 0,
                "translation_status": 1,
                "source_text": "室内カメラ",
                "translated_text": "開心",
                "storage_path": "TrsData.json",
                "file_name": "TrsData.json",
            },
            {
                "text_index": 3,
                "text_classification": 0,
                "translation_status": 0,
                "source_text": "111111111",
                "translated_text": "歷史",
                "storage_path": "DEBUG Folder\\Replace the original text.json",
                "file_name": "Replace the original text.json",
            },
            {
                "text_index": 4,
                "text_classification": 0,
                "translation_status": 0,
                "source_text": "222222222",
                "translated_text": "无",
                "storage_path": "DEBUG Folder\\Replace the original text.json",
                "file_name": "Replace the original text.json",
            },
        ]

        # 确定使用的转换器
        if target_language == "简中":
            cc = opencc.OpenCC("t2s")  # 创建OpenCC对象，使用t2s参数表示繁体字转简体字
        elif target_language == "繁中":
            cc = opencc.OpenCC("s2t")

        # 存储结果的列表
        converted_list = []

        # 遍历缓存数据
        for item in cache_list:
            translation_status = item.get("translation_status", 0)
            translated_text = item.get("translated_text", "")

            # 如果'translation_status'为1，进行转换
            if translation_status == 1:
                converted_text = cc.convert(translated_text)
                item_copy = item.copy()  # 防止修改原始数据
                item_copy["translated_text"] = converted_text
                converted_list.append(item_copy)
            else:
                converted_list.append(item)

        return converted_list

    # 输出json文件
    def output_json_file(self, cache_data, output_path):
        # 缓存数据结构示例
        ex_cache_data = [
            {"project_type": "Mtool"},
            {
                "text_index": 1,
                "text_classification": 0,
                "translation_status": 0,
                "source_text": "しこトラ！",
                "translated_text": "无",
                "storage_path": "TrsData.json",
                "file_name": "TrsData.json",
            },
            {
                "text_index": 2,
                "text_classification": 0,
                "translation_status": 0,
                "source_text": "室内カメラ",
                "translated_text": "无",
                "storage_path": "TrsData.json",
                "file_name": "TrsData.json",
            },
            {
                "text_index": 3,
                "text_classification": 0,
                "translation_status": 0,
                "source_text": "111111111",
                "translated_text": "无",
                "storage_path": "DEBUG Folder\\Replace the original text.json",
                "file_name": "Replace the original text.json",
            },
            {
                "text_index": 4,
                "text_classification": 0,
                "translation_status": 0,
                "source_text": "222222222",
                "translated_text": "无",
                "storage_path": "DEBUG Folder\\Replace the original text.json",
                "file_name": "Replace the original text.json",
            },
        ]

        # 中间存储字典格式示例
        ex_path_dict = {
            "D:\\DEBUG Folder\\Replace the original text.json": {
                "translation_status": 1,
                "Source Text": "しこトラ！",
                "Translated Text": "しこトラ！",
            },
            "D:\\DEBUG Folder\\DEBUG Folder\\Replace the original text.json": {
                "translation_status": 0,
                "Source Text": "しこトラ！",
                "Translated Text": "しこトラ！",
            },
        }

        # 输出文件格式示例
        ex_output = {
            "しこトラ！": "xxxx",
            "室内カメラ": "yyyyy",
            "111111111": "无3",
            "222222222": "无4",
        }

        # 创建中间存储字典，这个存储已经翻译的内容
        path_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if "storage_path" not in item:
                continue

            # 获取相对文件路径
            storage_path = item["storage_path"]
            # 获取文件名
            file_name = item["file_name"]

            if file_name != storage_path:
                # 构建文件输出路径
                file_path = f"{output_path}/{storage_path}"
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f"{output_path}/{storage_path}"

            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in path_dict:

                text = {
                    "translation_status": item["translation_status"],
                    "source_text": item["source_text"],
                    "translated_text": item["translated_text"],
                }
                path_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                text = {
                    "translation_status": item["translation_status"],
                    "source_text": item["source_text"],
                    "translated_text": item["translated_text"],
                }
                path_dict[file_path] = [text]

        # 遍历 path_dict，并将内容写入文件
        for file_path, content_list in path_dict.items():

            # 提取文件路径的文件夹路径和文件名
            folder_path, old_filename = os.path.split(file_path)

            # 创建已翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_translated = (
                    old_filename.replace(".json", "") + "_translated.json"
                )
            else:
                file_name_translated = old_filename + "_translated.json"
            file_path_translated = os.path.join(folder_path, file_name_translated)

            # 创建未翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_untranslated = (
                    old_filename.replace(".json", "") + "_untranslated.json"
                )
            else:
                file_name_untranslated = old_filename + "_untranslated.json"
            file_path_untranslated = os.path.join(folder_path, file_name_untranslated)

            # 存储已经翻译的文本
            output_file = {}

            # 存储未翻译的文本
            output_file2 = {}

            # 转换中间字典的格式为最终输出格式
            for content in content_list:
                # 如果这个本已经翻译了，存放对应的文件中
                if content["translation_status"] == 1:
                    output_file[content["source_text"]] = content["translated_text"]
                # 如果这个文本没有翻译或者正在翻译
                elif (
                    content["translation_status"] == 0
                    or content["translation_status"] == 2
                ):
                    output_file2[content["source_text"]] = content["source_text"]

            # 输出已经翻译的文件
            with open(file_path_translated, "w", encoding="utf-8") as file:
                json.dump(output_file, file, ensure_ascii=False, indent=4)

            # 输出未翻译的内容
            if output_file2:
                with open(file_path_untranslated, "w", encoding="utf-8") as file:
                    json.dump(output_file2, file, ensure_ascii=False, indent=4)

    # 输出表格文件
    def output_excel_file(self, cache_data, output_path):
        # 缓存数据结构示例
        ex_cache_data = [
            {"project_type": "T++"},
            {
                "text_index": 1,
                "text_classification": 0,
                "translation_status": 1,
                "source_text": "しこトラ！",
                "translated_text": "无",
                "storage_path": "TrsData.xlsx",
                "file_name": "TrsData.xlsx",
                "row_index": 2,
            },
            {
                "text_index": 2,
                "text_classification": 0,
                "translation_status": 0,
                "source_text": "室内カメラ",
                "translated_text": "无",
                "storage_path": "TrsData.xlsx",
                "file_name": "TrsData.xlsx",
                "row_index": 3,
            },
            {
                "text_index": 3,
                "text_classification": 0,
                "translation_status": 0,
                "source_text": "草草草草",
                "translated_text": "11111",
                "storage_path": "DEBUG Folder\\text.xlsx",
                "file_name": "text.xlsx",
                "row_index": 3,
            },
            {
                "text_index": 4,
                "text_classification": 0,
                "translation_status": 1,
                "source_text": "室内カメラ",
                "translated_text": "22222",
                "storage_path": "DEBUG Folder\\text.xlsx",
                "file_name": "text.xlsx",
                "row_index": 4,
            },
        ]

        # 创建一个字典，用于存储翻译数据
        translations_by_path = {}

        # 遍历缓存数据
        for item in cache_data:
            if "storage_path" in item:
                path = item["storage_path"]

                # 如果路径不存在，创建文件夹
                folder_path = os.path.join(output_path, os.path.dirname(path))
                os.makedirs(folder_path, exist_ok=True)

                # 提取信息
                source_text = item.get("source_text", "")
                translated_text = item.get("translated_text", "")
                row_index = item.get("row_index", "")
                translation_status = item.get("translation_status", "")

                # 构造字典
                translation_dict = {
                    "translation_status": translation_status,
                    "Source Text": source_text,
                    "Translated Text": translated_text,
                    "row_index": row_index,
                }

                # 将字典添加到对应路径的列表中
                if path in translations_by_path:
                    translations_by_path[path].append(translation_dict)
                else:
                    translations_by_path[path] = [translation_dict]

        # 遍历字典，将数据写入 Excel 文件
        for path, translations_list in translations_by_path.items():
            file_path = os.path.join(output_path, path)

            # 创建一个工作簿
            wb = Workbook()

            # 选择默认的活动工作表
            ws = wb.active

            # 添加表头
            ws.append(
                [
                    "Original Text",
                    "Initial",
                    "Machine translation",
                    "Better translation",
                    "Best translation",
                ]
            )

            # 将数据写入工作表
            for translation_dict in translations_list:
                row_index = translation_dict["row_index"]
                translation_status = translation_dict["translation_status"]

                # 如果是已经翻译文本，则写入原文与译文
                if translation_status == 1:
                    ws.cell(row=row_index, column=1).value = translation_dict[
                        "Source Text"
                    ]
                    ws.cell(row=row_index, column=2).value = translation_dict[
                        "Translated Text"
                    ]

                # 如果是未翻译或不需要翻译文本，则写入原文
                else:
                    ws.cell(row=row_index, column=1).value = translation_dict[
                        "Source Text"
                    ]

            # 保存工作簿
            wb.save(file_path)

    # 输出缓存文件
    def output_cache_file(self, cache_data, output_path):
        # 复制缓存数据到新变量
        modified_cache_data = copy.deepcopy(cache_data)

        # 修改新变量的元素中的'translation_status'
        for item in modified_cache_data:
            if "translation_status" in item and item["translation_status"] == 2:
                item["translation_status"] = 0

        # 输出为JSON文件
        with open(
            os.path.join(output_path, "AinieeCacheData.json"), "w", encoding="utf-8"
        ) as f:
            json.dump(modified_cache_data, f, ensure_ascii=False, indent=4)

    # 输出已经翻译文件
    def output_translated_content(self, cache_data, output_path):
        # 复制缓存数据到新变量
        new_cache_data = copy.deepcopy(cache_data)

        # 提取项目列表
        if new_cache_data[0]["project_type"] == "Mtool":
            File_Outputter.output_json_file(self, new_cache_data, output_path)
        else:
            File_Outputter.output_excel_file(self, new_cache_data, output_path)
